#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Bygger regnearket med fanerne Kandidater, Fravalgte og Metode og forbehold.

    python3 build_workbook.py data.json "Screening.xlsx"

Kolonnenavnene på fane 1 er faste, fordi det er dem, modtageren filtrerer og sorterer
efter. Beløb formateres med enheden i cellen, så overskriften kan stå uændret.
"""
import json, subprocess, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
NAVY, LIGHT, AMBER, GREY = "1F3864", "EDF1F7", "FFF2CC", "595959"
MIO = '#,##0.0" mio."'
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

KOL_KANDIDAT = ["Selskabsnavn", "CVR", "Segment", "By", "Omsætning", "Bruttofortjeneste",
                "Regnskabsår", "Antal ansatte", "Egenkapital", "Ejerforhold",
                "Kort beskrivelse af forretningen", "Vurdering af fit", "Usikkerheder", "Kilde",
                "Hjemmeside"]
KOL_FRAVALGT = ["Selskabsnavn", "CVR", "Segment", "Nøgletal (seneste offentliggjorte)",
                "Begrundelse for fravalg", "Kilde"]


def style_header(ws, n, row=1):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


def dato(iso):
    """2026-04-20 -> 20.04.2026; alt andet uændret."""
    try:
        y, m, d = str(iso).split("-")
        return f"{d}.{m}.{y}"
    except ValueError:
        return str(iso)


def begrundelse_med_ejerskifte(f):
    """Skriver et 'ejerskifte'-objekt ind under begrundelsen, så begge datoer og deres
    kilder står i regnearket — annoncering og closing hver for sig."""
    tekst = f.get("begrundelse") or ""
    e = f.get("ejerskifte")
    if not isinstance(e, dict):
        return tekst
    dele = []
    if e.get("annonceret"):
        dele.append(f"annonceret {dato(e['annonceret'])}"
                    + (f" ({e['kilde_annoncering']})" if e.get("kilde_annoncering") else ""))
    if e.get("closing"):
        dele.append(f"closing {dato(e['closing'])}"
                    + (f" ({e['kilde_closing']})" if e.get("kilde_closing") else ""))
    if e.get("koeber"):
        dele.append(f"køber: {e['koeber']}")
    if e.get("saelgers_andel_efter"):
        dele.append(f"sælgers andel efter: {e['saelgers_andel_efter']}")
    return tekst + ("\n\nEjerskifte: " + "; ".join(dele) if dele else "")


def num(ws, r, c, v, fmt):
    cell = ws.cell(r, c, v)
    cell.number_format = fmt
    return cell


def build(d, ud):
    wb = Workbook()

    # ---------------- Kandidater ----------------
    ws = wb.active
    ws.title = "Kandidater"
    ws.append(KOL_KANDIDAT)
    style_header(ws, len(KOL_KANDIDAT))
    for i, k in enumerate(d["kandidater"]):
        r = i + 2
        ws.cell(r, 1, k["navn"]); ws.cell(r, 2, str(k["cvr"]))
        ws.cell(r, 3, k["segment"]); ws.cell(r, 4, k["by"])
        if k.get("omsaetning") is None:
            ws.cell(r, 5, "ikke oplyst")
        else:
            num(ws, r, 5, k["omsaetning"], MIO)
        num(ws, r, 6, k["bruttofortjeneste"], MIO)
        ws.cell(r, 7, k["regnskabsaar"])
        num(ws, r, 8, k["ansatte"], "#,##0")
        num(ws, r, 9, k["egenkapital"], MIO)
        for col, felt in [(10, "ejerforhold"), (11, "beskrivelse"), (12, "fit"),
                          (13, "usikkerheder"), (14, "kilde")]:
            ws.cell(r, col, k[felt])
        ws.cell(r, 15, k.get("hjemmeside") or "ikke undersøgt")

    n = len(d["kandidater"]) + 1
    for r in range(2, n + 1):
        for c in range(1, len(KOL_KANDIDAT) + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(r, 1).font = Font(name=FONT, size=10, bold=True)
        if ws.cell(r, 5).value == "ikke oplyst":
            ws.cell(r, 5).font = Font(name=FONT, size=10, italic=True, color=GREY)
        if str(ws.cell(r, 13).value).startswith(("OMSÆTNING ESTIMERET", "EJERFORHOLD")):
            ws.cell(r, 13).fill = PatternFill("solid", fgColor=AMBER)
        if r % 2 == 0:
            for c in range(1, len(KOL_KANDIDAT) + 1):
                if ws.cell(r, c).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
        ws.row_dimensions[r].height = 150
    for i, w in enumerate([30, 11, 30, 22, 14, 16, 20, 11, 14, 46, 46, 46, 56, 46, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(KOL_KANDIDAT))}{n}"

    # ---------------- Fravalgte ----------------
    ws2 = wb.create_sheet("Fravalgte")
    ws2.append(KOL_FRAVALGT)
    style_header(ws2, len(KOL_FRAVALGT))
    for f in d.get("fravalgte", []):
        ws2.append([f.get("navn"), str(f.get("cvr", "—")), f.get("segment"),
                    f.get("noegletal", "ikke opgjort"), begrundelse_med_ejerskifte(f), f.get("kilde")])
    n2 = len(d.get("fravalgte", [])) + 1
    for r in range(2, n2 + 1):
        for c in range(1, len(KOL_FRAVALGT) + 1):
            cell = ws2.cell(r, c)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
        ws2.cell(r, 1).font = Font(name=FONT, size=10, bold=True)
        ws2.row_dimensions[r].height = 118
    for i, w in enumerate([30, 11, 32, 40, 82, 44], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:F{n2}"

    # ---------------- Metode og forbehold ----------------
    ws3 = wb.create_sheet("Metode og forbehold")
    ws3.column_dimensions["A"].width = 3
    ws3.column_dimensions["B"].width = 34
    ws3.column_dimensions["C"].width = 118
    r = 1
    for kind, text in d.get("metode", []):
        if kind == "h1":
            c = ws3.cell(r, 2, text); c.font = Font(name=FONT, size=16, bold=True, color=NAVY)
            ws3.row_dimensions[r].height = 26; r += 2
        elif kind == "h2":
            c = ws3.cell(r, 2, text); c.font = Font(name=FONT, size=12, bold=True, color=NAVY)
            ws3.row_dimensions[r].height = 24; r += 1
        elif kind == "p":
            c = ws3.cell(r, 2, text); c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            ws3.row_dimensions[r].height = max(15, 13 * (len(text) // 130 + 1)); r += 1
        elif kind == "b":
            ws3.cell(r, 2, "•").font = Font(name=FONT, size=10, color=NAVY)
            ws3.cell(r, 2).alignment = Alignment(horizontal="right", vertical="top")
            c = ws3.cell(r, 3, text); c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws3.row_dimensions[r].height = max(15, 13 * (len(text) // 105 + 1)); r += 1

    oev = d.get("oevrige_fravalgte", [])
    if oev:
        r += 1
        for col, txt in [(2, "Årsag"), (3, "Selskaber")]:
            c = ws3.cell(r, col, txt)
            c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY)
        ws3.row_dimensions[r].height = 18; r += 1
        for o in oev:
            a = ws3.cell(r, 2, f"{o['aarsag']} ({o['antal']})")
            a.font = Font(name=FONT, size=10, bold=True)
            a.alignment = Alignment(wrap_text=True, vertical="top"); a.border = BORDER
            b = ws3.cell(r, 3, o.get("selskaber", ""))
            b.font = Font(name=FONT, size=10)
            b.alignment = Alignment(wrap_text=True, vertical="top"); b.border = BORDER
            ws3.row_dimensions[r].height = max(30, 13 * (len(o.get("selskaber", "")) // 105 + 1))
            r += 1

    r += 2
    ws3.cell(r, 2, "Nøgletal for screeningen").font = Font(name=FONT, size=12, bold=True, color=NAVY)
    r += 1
    raekker = [
        ("Antal kandidater i fanen 'Kandidater'", "=COUNTA(Kandidater!A2:A400)", "#,##0"),
        ("Heraf uden oplyst omsætning (skøn anvendt)", '=COUNTIF(Kandidater!E2:E400,"ikke oplyst")', "#,##0"),
        ("Heraf med oplyst omsætning i årsrapporten",
         '=COUNTA(Kandidater!A2:A400)-COUNTIF(Kandidater!E2:E400,"ikke oplyst")', "#,##0"),
        ("Samlet antal ansatte i kandidatlisten", "=SUM(Kandidater!H2:H400)", "#,##0"),
        ("Samlet egenkapital i kandidatlisten (mio. DKK)", "=SUM(Kandidater!I2:I400)", "#,##0.0"),
        ("Antal selskaber i fanen 'Fravalgte'", "=COUNTA(Fravalgte!A2:A400)", "#,##0"),
    ]
    for label, formel, fmt in raekker:
        ws3.cell(r, 2, label).font = Font(name=FONT, size=10)
        ws3.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="center")
        c = ws3.cell(r, 3, formel); c.number_format = fmt
        c.font = Font(name=FONT, size=10, bold=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws3.row_dimensions[r].height = 16; r += 1
    if d.get("meta", {}).get("antal_vurderet"):
        ws3.cell(r, 2, "Selskaber vurderet i alt (optalt manuelt fra researchloggen)").font = Font(name=FONT, size=10)
        c = ws3.cell(r, 3, d["meta"]["antal_vurderet"])
        c.font = Font(name=FONT, size=10, color="0000FF"); c.number_format = '#,##0'
        ws3.row_dimensions[r].height = 16; r += 1
    r += 1
    ws3.cell(r, 2, "Blå skrift = tal indtastet manuelt. Sort skrift = beregnet med formel ud fra "
                   "de to øvrige faner. Beløb er i mio. DKK. Kilde til hvert enkelt selskabs tal "
                   "står i kolonnen Kilde.").font = Font(name=FONT, size=9, italic=True, color=GREY)
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws3.row_dimensions[r].height = 28
    ws3.sheet_view.showGridLines = False

    wb.save(ud)


def recalc(sti):
    """Får LibreOffice til at beregne formlerne, så cachede værdier findes i filen."""
    for kandidat in ("soffice", "libreoffice"):
        try:
            subprocess.run([kandidat, "--headless", "--convert-to", "xlsx", "--outdir",
                            os.path.dirname(os.path.abspath(sti)) or ".", sti],
                           capture_output=True, timeout=120, check=True)
            return True
        except (FileNotFoundError, subprocess.SubprocessError):
            continue
    return False


def main():
    if len(sys.argv) < 3:
        print("brug: build_workbook.py <data.json> <ud.xlsx>"); sys.exit(2)
    with open(sys.argv[1], encoding="utf-8") as f:
        d = json.load(f)
    build(d, sys.argv[2])
    ok = recalc(sys.argv[2])
    print(f"skrevet: {sys.argv[2]}  ({len(d['kandidater'])} kandidater, "
          f"{len(d.get('fravalgte', []))} fravalgte)")
    if not ok:
        print("bemærk: LibreOffice blev ikke fundet, så formlerne i metodefanen har ingen "
              "cachede værdier endnu. Excel beregner dem, når filen åbnes.")


if __name__ == "__main__":
    main()
